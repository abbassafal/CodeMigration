import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from db_config import (
    get_mssql_connection, 
    get_postgres_connection,
    print_header,
    get_output_path
)

print_header("Simple Table Comparison - Side by Side")

# Connect to databases
print("\n[1/3] Connecting to MSSQL...")
mssql_conn = get_mssql_connection()
print("✓ MSSQL Connected")

print("\n[2/3] Connecting to PostgreSQL...")
pg_conn = get_postgres_connection()
print("✓ PostgreSQL Connected")

# Get table names from user
print("\n" + "=" * 80)
mssql_table = input("Enter MSSQL table name: ")
pg_table = input("Enter PostgreSQL table name: ")
print("=" * 80)

print(f"\nComparing:")
print(f"  MSSQL: {mssql_table}")
print(f"  PostgreSQL: {pg_table}")

try:
    # Get MSSQL columns with constraints and foreign keys
    print("\n[3/3] Fetching table structures...")
    
    # Use direct cursor approach for better control
    cursor = mssql_conn.cursor()
    cursor.execute("""
        SELECT 
            c.COLUMN_NAME,
            c.DATA_TYPE,
            CASE WHEN c.IS_NULLABLE = 'NO' THEN 'NOT NULL' ELSE 'NULLABLE' END AS NULLABLE,
            ISNULL(fk.FK_INFO, '') AS FOREIGN_KEY
        FROM INFORMATION_SCHEMA.COLUMNS c
        LEFT JOIN (
            SELECT 
                fk.parent_object_id,
                COL_NAME(fkc.parent_object_id, fkc.parent_column_id) AS column_name,
                OBJECT_NAME(fk.referenced_object_id) + '(' + COL_NAME(fkc.referenced_object_id, fkc.referenced_column_id) + ')' AS FK_INFO
            FROM sys.foreign_keys fk
            INNER JOIN sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
        ) fk ON c.TABLE_NAME = OBJECT_NAME(fk.parent_object_id) AND c.COLUMN_NAME = fk.column_name
        WHERE c.TABLE_NAME = ? AND c.TABLE_SCHEMA = 'dbo'
        ORDER BY c.ORDINAL_POSITION
    """, (mssql_table,))
    
    mssql_rows = cursor.fetchall()
    cursor.close()
    
    if len(mssql_rows) == 0:
        print(f"✗ Table '{mssql_table}' not found in MSSQL!")
        mssql_conn.close()
        pg_conn.close()
        input("\nPress Enter to exit...")
        exit()
    
    # Convert to DataFrame - handle the cursor result properly
    mssql_data = []
    for row in mssql_rows:
        mssql_data.append({
            'COLUMN_NAME': row[0], 
            'DATA_TYPE': row[1],
            'NULLABLE': row[2],
            'FOREIGN_KEY': row[3] if row[3] else ''
        })
    df_mssql = pd.DataFrame(mssql_data)
    
    print(f"✓ MSSQL: {len(df_mssql)} columns")
    
    # Get PostgreSQL columns using cursor approach with constraints and foreign keys
    cursor = pg_conn.cursor()
    cursor.execute("""
        SELECT 
            c.column_name,
            c.data_type,
            CASE WHEN c.is_nullable = 'NO' THEN 'NOT NULL' ELSE 'NULLABLE' END AS nullable,
            COALESCE(
                (SELECT ccu.table_name || '(' || ccu.column_name || ')'
                 FROM information_schema.table_constraints tc
                 JOIN information_schema.key_column_usage kcu 
                   ON tc.constraint_name = kcu.constraint_name
                   AND tc.table_schema = kcu.table_schema
                 JOIN information_schema.constraint_column_usage ccu 
                   ON ccu.constraint_name = tc.constraint_name
                   AND ccu.table_schema = tc.table_schema
                 WHERE tc.constraint_type = 'FOREIGN KEY'
                   AND tc.table_name = %s
                   AND kcu.column_name = c.column_name
                   AND tc.table_schema = 'public'
                 LIMIT 1), ''
            ) AS foreign_key
        FROM information_schema.columns c
        WHERE c.table_name = %s AND c.table_schema = 'public'
        ORDER BY c.ordinal_position
    """, (pg_table, pg_table))
    
    pg_rows = cursor.fetchall()
    cursor.close()
    
    if len(pg_rows) == 0:
        print(f"✗ Table '{pg_table}' not found in PostgreSQL!")
        mssql_conn.close()
        pg_conn.close()
        input("\nPress Enter to exit...")
        exit()
    
    # Convert to DataFrame - handle the cursor result properly
    pg_data = []
    for row in pg_rows:
        pg_data.append({
            'column_name': row[0], 
            'data_type': row[1],
            'nullable': row[2],
            'foreign_key': row[3] if row[3] else ''
        })
    df_pg = pd.DataFrame(pg_data)
    
    print(f"✓ PostgreSQL: {len(df_pg)} columns")
    
    # Create side-by-side comparison data
    max_rows = max(len(df_mssql), len(df_pg))
    
    comparison_data = []
    
    for i in range(max_rows):
        row_data = {}
        
        # MSSQL columns
        if i < len(df_mssql):
            row_data['MSSQL_COLUMN_NAME'] = df_mssql.iloc[i]['COLUMN_NAME']
            row_data['MSSQL_DATA_TYPE'] = df_mssql.iloc[i]['DATA_TYPE']
            row_data['MSSQL_NULLABLE'] = df_mssql.iloc[i]['NULLABLE']
            row_data['MSSQL_FOREIGN_KEY'] = df_mssql.iloc[i]['FOREIGN_KEY']
        else:
            row_data['MSSQL_COLUMN_NAME'] = ''
            row_data['MSSQL_DATA_TYPE'] = ''
            row_data['MSSQL_NULLABLE'] = ''
            row_data['MSSQL_FOREIGN_KEY'] = ''
        
        # Empty separator column
        row_data['SEPARATOR'] = ''
        
        # PostgreSQL columns
        if i < len(df_pg):
            row_data['PG_column_name'] = df_pg.iloc[i]['column_name']
            row_data['PG_data_type'] = df_pg.iloc[i]['data_type']
            row_data['PG_nullable'] = df_pg.iloc[i]['nullable']
            row_data['PG_foreign_key'] = df_pg.iloc[i]['foreign_key']
        else:
            row_data['PG_column_name'] = ''
            row_data['PG_data_type'] = ''
            row_data['PG_nullable'] = ''
            row_data['PG_foreign_key'] = ''
        
        comparison_data.append(row_data)
    
    df_comparison = pd.DataFrame(comparison_data)
    
    # Save to Excel
    output_file = get_output_path(f"Compare_{mssql_table}_vs_{pg_table}.xlsx")
    
    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_comparison.to_excel(writer, sheet_name='Comparison', index=False, startrow=1)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Comparison']
        
        # Add table names in first row
        worksheet['A1'] = f'mssql - {mssql_table}'
        worksheet['F1'] = f'postgres - {pg_table}'
        
        # Merge cells for headers
        worksheet.merge_cells('A1:D1')
        worksheet.merge_cells('F1:I1')
        
        # Style the header row (table names)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        
        for cell in ['A1', 'F1']:
            worksheet[cell].fill = header_fill
            worksheet[cell].font = header_font
            worksheet[cell].alignment = center_align
        
        # Style column headers (COLUMN_NAME, DATA_TYPE, etc.)
        col_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        col_header_font = Font(bold=True, size=11)
        
        # Rename column headers
        worksheet['A2'] = 'COLUMN_NAME'
        worksheet['B2'] = 'DATA_TYPE'
        worksheet['C2'] = 'NULLABLE'
        worksheet['D2'] = 'FOREIGN_KEY'
        worksheet['E2'] = ''
        worksheet['F2'] = 'column_name'
        worksheet['G2'] = 'data_type'
        worksheet['H2'] = 'nullable'
        worksheet['I2'] = 'foreign_key'
        
        for cell in ['A2', 'B2', 'C2', 'D2', 'F2', 'G2', 'H2', 'I2']:
            worksheet[cell].fill = col_header_fill
            worksheet[cell].font = col_header_font
            worksheet[cell].alignment = center_align
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 18
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 3
        worksheet.column_dimensions['F'].width = 28
        worksheet.column_dimensions['G'].width = 25
        worksheet.column_dimensions['H'].width = 12
        worksheet.column_dimensions['I'].width = 35
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=max_rows+2, min_col=1, max_col=9):
            for cell in row:
                if cell.column != 5:  # Skip separator column E
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Make separator column E gray
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in range(1, max_rows + 3):
            worksheet[f'E{row}'].fill = gray_fill
        
        # Highlight NOT NULL cells in light green
        not_null_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for row in range(3, max_rows + 3):  # Start from row 3 (after headers)
            # MSSQL NOT NULL
            if worksheet[f'C{row}'].value == 'NOT NULL':
                worksheet[f'C{row}'].fill = not_null_fill
            # PostgreSQL NOT NULL
            if worksheet[f'H{row}'].value == 'NOT NULL':
                worksheet[f'H{row}'].fill = not_null_fill
        
        # Highlight cells with foreign keys in light blue
        fk_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for row in range(3, max_rows + 3):
            # MSSQL FK
            if worksheet[f'D{row}'].value and str(worksheet[f'D{row}'].value).strip():
                worksheet[f'D{row}'].fill = fk_fill
            # PostgreSQL FK
            if worksheet[f'I{row}'].value and str(worksheet[f'I{row}'].value).strip():
                worksheet[f'I{row}'].fill = fk_fill
    
    print("\n" + "=" * 80)
    print("✓ COMPARISON COMPLETE!")
    print("=" * 80)
    print(f"\n✓ Enhanced comparison saved to:")
    print(f"  {output_file}")
    print(f"\nFormat:")
    print(f"  - MSSQL columns on the left (COLUMN_NAME, DATA_TYPE, NULLABLE, FOREIGN_KEY)")
    print(f"  - PostgreSQL columns on the right (column_name, data_type, nullable, foreign_key)")
    print(f"  - NOT NULL columns highlighted in light green")
    print(f"  - Foreign key columns highlighted in light blue")
    print(f"  - Clean, formatted Excel sheet with all constraints visible")

except Exception as e:
    print(f"\n✗ Comparison Failed!")
    print(f"Error: {str(e)}")
    import traceback
    traceback.print_exc()

finally:
    mssql_conn.close()
    pg_conn.close()
    print("\n" + "=" * 80)
    print("Connections closed")
    print("=" * 80)

input("\nPress Enter to exit...")




